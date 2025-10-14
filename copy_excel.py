# copy_excel.py
import os
import re
import openpyxl
from copy import copy
from contextlib import contextmanager

def get_file_extension(file_path):
    """Возвращает расширение файла (.xlsx)"""
    return '.xlsx'

def get_column_letter(col_idx):
    """Конвертирует индекс столбца в букву (1 -> A, 26 -> Z, 27 -> AA и т.д.)"""
    letters = []
    while col_idx > 0:
        col_idx, remainder = divmod(col_idx - 1, 26)
        letters.append(chr(65 + remainder))
    return ''.join(reversed(letters))

@contextmanager
def safe_workbook(file_path, read_only=False):
    """Контекстный менеджер для безопасной работы с файлами Excel."""
    wb = None
    try:
        wb = openpyxl.load_workbook(file_path, read_only=read_only)
        yield wb
    finally:
        if wb:
            try:
                wb.close()
            except:
                pass

def get_all_sheets_headers(file_path, max_scan_rows=10):
    """Анализирует все ВИДИМЫЕ листы в Excel-файле, возвращает заголовки для каждого."""
    try:
        with safe_workbook(file_path, read_only=True) as wb:
            sheet_results = {}
            
            for ws in wb.worksheets:
                # Игнорируем скрытые листы
                if ws.sheet_state != 'visible':
                    continue
                    
                max_non_empty = 0
                header_row = None
                header_row_idx = 0

                for row_idx, row in enumerate(ws.iter_rows(min_row=1, max_row=max_scan_rows), start=1):
                    non_empty_count = sum(1 for cell in row if cell.value is not None)
                    if non_empty_count > max_non_empty:
                        max_non_empty = non_empty_count
                        header_row = row
                        header_row_idx = row_idx

                if max_non_empty > 0:
                    headers = [cell.value for cell in header_row if cell.value is not None]
                    sheet_results[ws.title] = (headers, header_row_idx)
                else:
                    sheet_results[ws.title] = (None, None)
                    
            return sheet_results
    except Exception as e:
        raise ValueError(f"Error analyzing Excel: {str(e)}")

def analyze_column(file_path, valid_sheets, selected_column, filters=None):
    """Собирает уникальные значения из указанной колонки с учетом фильтров."""
    if filters is None:
        filters = {}
    try:
        with safe_workbook(file_path, read_only=True) as wb:
            categories = set()
            
            for sheet_name, (headers, row_idx) in valid_sheets.items():
                ws = wb[sheet_name]
                try:
                    col_index = headers.index(selected_column)
                except ValueError:
                    continue
                
                for row in ws.iter_rows(min_row=row_idx + 1, values_only=True):
                    if not validate_row(row, headers, row_idx, filters):
                        continue
                    
                    cell_value = row[col_index] if col_index < len(row) else None
                    if cell_value is not None and str(cell_value).strip() != "":
                        categories.add(str(cell_value).strip())
            
            return sorted(categories)
    except Exception as e:
        raise ValueError(f"Error analyzing data: {str(e)}")

def validate_row(row, headers, header_row_idx, filters):
    """Проверяет соответствие строки условиям фильтров."""
    if not filters:
        return True
    
    for col, value in filters.items():
        try:
            col_index = headers.index(col)
            cell_value = row[col_index] if col_index < len(row) else None
            if str(cell_value).strip() != str(value).strip():
                return False
        except ValueError:
            return False
    return True

def sanitize_filename(name):
    """Удаляет недопустимые символы из названия файла."""
    return re.sub(r'[\\/*?:"<>|]', '_', name)

def create_filtered_file(source, target, valid_sheets, filters):
    """Создаёт файл с фильтрацией по комбинации условий."""
    try:
        # Всегда сохраняем как .xlsx
        if target.lower().endswith('.xlsm'):
            target = target[:-5] + '.xlsx'
        
        with safe_workbook(source, read_only=False) as wb_source:
            wb_new = openpyxl.Workbook()
            wb_new.remove(wb_new.active)
            
            has_data = False  # Флаг наличия данных
            
            for sheet_name in wb_source.sheetnames:
                ws_source = wb_source[sheet_name]
                
                # Игнорируем скрытые листы
                if ws_source.sheet_state != 'visible':
                    continue
                
                ws_new = wb_new.create_sheet(title=sheet_name)
                
                # Копирование ширины столбцов
                if hasattr(ws_source, 'column_dimensions'):
                    for col_letter, dim in ws_source.column_dimensions.items():
                        try:
                            ws_new.column_dimensions[col_letter].width = dim.width
                        except:
                            pass
                
                # Копирование высоты строк
                if hasattr(ws_source, 'row_dimensions'):
                    for row_idx, dim in ws_source.row_dimensions.items():
                        try:
                            ws_new.row_dimensions[row_idx].height = dim.height
                        except:
                            pass
                
                # Копирование объединенных ячеек
                if hasattr(ws_source, 'merged_cells'):
                    for merged_cell in ws_source.merged_cells.ranges:
                        try:
                            ws_new.merge_cells(str(merged_cell))
                        except:
                            pass
                
                # Копирование условного форматирования
                if hasattr(ws_source, 'conditional_formatting'):
                    for cf in ws_source.conditional_formatting:
                        try:
                            # Пытаемся скопировать условное форматирование
                            ws_new.conditional_formatting.add(cf._range, cf)
                        except Exception as e:
                            # Пытаемся создать новое условное форматирование с сохранением свойств
                            try:
                                new_cf = type(cf)()
                                
                                # Копируем основные свойства
                                for attr in ['type', 'operator', 'formula', 'dxf', 'pivot', 'stopIfTrue', 'priority']:
                                    if hasattr(cf, attr):
                                        setattr(new_cf, attr, getattr(cf, attr))
                                
                                # Копируем цвета для правил с цветами
                                if hasattr(cf, 'colorScale'):
                                    if hasattr(cf.colorScale, 'cfvo'):
                                        new_cf.colorScale.cfvo = [copy(cfvo) for cfvo in cf.colorScale.cfvo]
                                    if hasattr(cf.colorScale, 'color'):
                                        new_cf.colorScale.color = [copy(color) for color in cf.colorScale.color]
                                
                                # Копируем данные для правил с данными
                                if hasattr(cf, 'dataBar'):
                                    if hasattr(cf.dataBar, 'cfvo'):
                                        new_cf.dataBar.cfvo = [copy(cfvo) for cfvo in cf.dataBar.cfvo]
                                    if hasattr(cf.dataBar, 'color'):
                                        new_cf.dataBar.color = copy(cf.dataBar.color)
                                
                                ws_new.conditional_formatting.add(cf._range, new_cf)
                            except Exception as e:
                                # Пытаемся скопировать минимально необходимые атрибуты
                                try:
                                    new_cf = openpyxl.formatting.rule.Rule(
                                        type=cf.type,
                                        dxf=cf.dxf,
                                        operator=cf.operator,
                                        formula=cf.formula
                                    )
                                    ws_new.conditional_formatting.add(cf._range, new_cf)
                                except:
                                    pass
                
                if sheet_name in valid_sheets:
                    headers, header_row_idx = valid_sheets[sheet_name]
                    
                    # 1. Технические строки выше таблицы
                    for row_idx in range(1, header_row_idx):
                        for col_idx in range(1, ws_source.max_column + 1):
                            try:
                                cell = ws_source.cell(row=row_idx, column=col_idx)
                                if cell.value is not None or cell.has_style:
                                    new_cell = ws_new.cell(row=row_idx, column=col_idx, value=cell.value)
                                    if cell.has_style:
                                        new_cell.style = cell.style
                                        new_cell.font = copy(cell.font)
                                        new_cell.border = copy(cell.border)
                                        new_cell.fill = copy(cell.fill)
                                        new_cell.alignment = copy(cell.alignment)
                                        new_cell.number_format = cell.number_format
                            except:
                                pass
                    
                    # 2. Заголовки
                    for col_idx in range(1, ws_source.max_column + 1):
                        try:
                            cell = ws_source.cell(row=header_row_idx, column=col_idx)
                            if cell.value is not None or cell.has_style:
                                new_cell = ws_new.cell(row=header_row_idx, column=col_idx, value=cell.value)
                                if cell.has_style:
                                    new_cell.style = cell.style
                                    new_cell.font = copy(cell.font)
                                    new_cell.border = copy(cell.border)
                                    new_cell.fill = copy(cell.fill)
                                    new_cell.alignment = copy(cell.alignment)
                                    new_cell.number_format = cell.number_format
                        except:
                            pass
                    
                    # 3. Фильтрация данных
                    new_row_idx = header_row_idx + 1
                    for row_idx in range(header_row_idx + 1, ws_source.max_row + 1):
                        try:
                            row = ws_source[row_idx]
                            if not validate_row([cell.value for cell in row], headers, header_row_idx, filters):
                                continue
                            
                            for col_idx in range(1, ws_source.max_column + 1):
                                try:
                                    source_cell = ws_source.cell(row=row_idx, column=col_idx)
                                    if source_cell.value is not None or source_cell.has_style:
                                        new_cell = ws_new.cell(row=new_row_idx, column=col_idx, value=source_cell.value)
                                        if source_cell.has_style:
                                            new_cell.style = source_cell.style
                                            new_cell.font = copy(source_cell.font)
                                            new_cell.border = copy(source_cell.border)
                                            new_cell.fill = copy(source_cell.fill)
                                            new_cell.alignment = copy(source_cell.alignment)
                                            new_cell.number_format = source_cell.number_format
                                except:
                                    pass
                            new_row_idx += 1
                        except:
                            pass
                    
                    # Проверка наличия данных
                    if new_row_idx > header_row_idx + 1:
                        has_data = True
                    else:
                        # Удаляем лист без данных
                        wb_new.remove(ws_new)
                        continue  # Переходим к следующему листу
                    
                    # 4. Активируем автофильтр только для строки заголовков
                    last_col_letter = get_column_letter(ws_source.max_column)
                    ws_new.auto_filter.ref = f"A{header_row_idx}:{last_col_letter}{header_row_idx}"
                else:
                    for row_idx in range(1, ws_source.max_row + 1):
                        for col_idx in range(1, ws_source.max_column + 1):
                            try:
                                cell = ws_source.cell(row=row_idx, column=col_idx)
                                if cell.value is not None or cell.has_style:
                                    new_cell = ws_new.cell(row=row_idx, column=col_idx, value=cell.value)
                                    if cell.has_style:
                                        new_cell.style = cell.style
                                        new_cell.font = copy(cell.font)
                                        new_cell.border = copy(cell.border)
                                        new_cell.fill = copy(cell.fill)
                                        new_cell.alignment = copy(cell.alignment)
                                        new_cell.number_format = cell.number_format
                            except:
                                pass
            
            if not has_data:
                # Нет данных - не сохраняем файл
                return None
            
            # Удаляем целевой файл, если он существует
            if os.path.exists(target):
                os.remove(target)
                
            # Сохраняем как .xlsx
            wb_new.save(target)
            return target
    except Exception as e:
        raise ValueError(f"Error during filtering: {str(e)}")

def get_all_combinations(source, valid_sheets, hierarchy_columns, filters=None, level=0):
    """Возвращает все возможные комбинации фильтров, включая частичные уровни."""
    if filters is None:
        filters = {}
    
    if level >= len(hierarchy_columns):
        return [filters.copy()]
    
    column = hierarchy_columns[level]
    categories = analyze_column(source, valid_sheets, column, filters)
    combinations = []
    
    # Добавляем комбинации для текущего уровня без добавления следующих уровней
    for category in categories:
        new_filters = filters.copy()
        new_filters[column] = category
        combinations.append(new_filters.copy())
    
    # Добавляем комбинации для следующих уровней
    for category in categories:
        new_filters = filters.copy()
        new_filters[column] = category
        combinations.extend(get_all_combinations(source, valid_sheets, hierarchy_columns, new_filters, level + 1))
    
    return combinations

def select_categories_sequentially(source, valid_sheets, hierarchy_columns):
    """Последовательно запрашивает выбор категорий у пользователя с отображением вариантов для каждой комбинации."""
    all_combinations = []
    
    def generate_combinations(level, current_filters, include_all=False):
        if level >= len(hierarchy_columns):
            all_combinations.append(current_filters.copy())
            return
        
        column = hierarchy_columns[level]
        categories = analyze_column(source, valid_sheets, column, current_filters)
        
        if not categories:
            return
        
        # Проверяем, является ли текущий уровень последним
        is_last_level = (level == len(hierarchy_columns) - 1)
        
        # Если это не первый уровень и есть предыдущие фильтры
        if level > 0:
            print(f"\nCurrent filters:")
            for col, value in current_filters.items():
                print(f"  - {col}: {value}")
            
            # Если не последний уровень, спрашиваем, хочет ли пользователь выбрать все комбинации
            if not is_last_level:
                while True:
                    all_comb = input(f"Include all categories from '{hierarchy_columns[level]}' for current filters? (y/n): ").strip().lower()
                    
                    if all_comb == 'y':
                        # Анализируем все возможные комбинации
                        for category in categories:
                            new_filters = current_filters.copy()
                            new_filters[column] = category
                            generate_combinations(level + 1, new_filters, True)
                        return
                    elif all_comb == 'n':
                        break
                    else:
                        print("Please enter 'y' or 'n'")
        
        # Выводим доступные категории с номерами
        print(f"\nAvailable categories for column '{column}':")
        for i, cat in enumerate(categories, 1):
            print(f"  {i}. {cat}")
        print("  b. Назад")
        print("  c. Отмена")
        
        # Запрашиваем выбор
        while True:
            selection = input(f"Enter categories for '{column}' (comma-separated numbers, 'all', 'b' for back, 'c' for cancel): ").strip()
            
            # Обработка специальных команд
            if selection.lower() in ["c", "cancel", "отмена"]:
                print("Operation cancelled by user")
                return
            
            if selection.lower() in ["b", "back", "назад"]:
                return
            
            # Обработка "all" - выбираем все категории
            if selection.lower() == "all":
                selected_categories = categories
                break
            
            # Обработка номеров
            user_categories = []
            invalid_inputs = []
            for item in selection.split(","):
                item = item.strip()
                if item.isdigit():
                    idx = int(item) - 1
                    if 0 <= idx < len(categories):
                        user_categories.append(categories[idx])
                    else:
                        invalid_inputs.append(item)
                else:
                    user_categories.append(item)
            
            # Проверка валидности
            invalid_categories = [cat for cat in user_categories if cat not in categories]
            if invalid_categories or invalid_inputs:
                invalid_list = invalid_categories + invalid_inputs
                print(f"Error: Invalid categories: {', '.join(invalid_list)}")
                continue
            
            selected_categories = user_categories
            break
        
        # Обрабатываем выбор
        for category in selected_categories:
            new_filters = current_filters.copy()
            new_filters[column] = category
            
            # Если пользователь выбрал "all" для предыдущего уровня
            if include_all and level > 0:
                generate_combinations(level + 1, new_filters, True)
            else:
                generate_combinations(level + 1, new_filters)
    
    # Начинаем генерацию комбинаций с первого уровня
    generate_combinations(0, {})
    
    # Добавляем частичные уровни фильтрации
    final_combinations = []
    for filters in all_combinations:
        # Добавляем фильтры всех подуровней
        for i in range(len(hierarchy_columns)):
            partial_filter = {}
            for j in range(i + 1):
                if hierarchy_columns[j] in filters:
                    partial_filter[hierarchy_columns[j]] = filters[hierarchy_columns[j]]
            final_combinations.append(partial_filter)
    
    # Удаляем дубликаты частичных фильтров
    unique_combinations = []
    seen = set()
    for filters in final_combinations:
        filter_tuple = tuple(sorted(filters.items()))
        if filter_tuple not in seen:
            seen.add(filter_tuple)
            unique_combinations.append(filters)
    
    return unique_combinations

def process_file():
    """Обрабатывает один файл: выбор файла, директории, колонок, категорий, создание файлов."""
    print("\n=== Copy Excel File ===")
    print("To cancel the operation, press Ctrl+C at any time")
    
    try:
        # Шаг 0: Выбор исходного файла
        while True:
            source = input("\nEnter full path to source Excel file: ").strip('"')
            if source.lower() == "cancel":
                print("Operation cancelled by user")
                return False
            
            if os.path.exists(source) and os.path.isfile(source):
                # Проверка формата файла
                if not (source.lower().endswith('.xlsx') or source.lower().endswith('.xlsm')):
                    print("Error: File must have .xlsx or .xlsm extension")
                    continue
                break
            print(f"Error: Source file not found or is not a file: {source}")
        
        # Шаг 1: Выбор целевой директории
        while True:
            destination = input("Enter target directory path: ").strip('"')
            if destination.lower() == "cancel":
                print("Operation cancelled by user")
                return False
            
            if os.path.isdir(destination):
                break
            print(f"Error: Target directory does not exist: {destination}")
        
        # Анализ Excel: заголовки во всех листах
        sheet_headers = get_all_sheets_headers(source)
        valid_sheets = {sheet: data for sheet, data in sheet_headers.items() if data[0] is not None}
        
        if not valid_sheets:
            print("Error: No headers found in any sheet")
            return False
        
        # Поиск пересечения заголовков
        all_headers = [set(headers) for headers, _ in valid_sheets.values()]
        common_headers = set.intersection(*all_headers) if all_headers else set()
        
        if not common_headers:
            print("\nWarning: No common headers found between sheets")
            return False
        
        # Шаг 2: Выбор колонок для фильтрации
        print("\nAvailable columns for filtering:")
        common_headers_list = list(common_headers)
        for i, col in enumerate(common_headers_list, 1):
            print(f"  {i}. {col}")
        print("  b. Назад")
        print("  c. Отмена")
        
        while True:
            columns_input = input("Enter columns for filtering (comma-separated numbers or names): ").strip()
            
            if columns_input.lower() in ["c", "cancel", "отмена"]:
                print("Operation cancelled by user")
                return False
            
            if columns_input.lower() in ["b", "back", "назад"]:
                return False  # Возврат к началу
            
            # Обработка номеров колонок
            hierarchy_columns = []
            invalid_inputs = []
            for item in columns_input.split(","):
                item = item.strip()
                if item.isdigit():
                    idx = int(item) - 1
                    if 0 <= idx < len(common_headers_list):
                        hierarchy_columns.append(common_headers_list[idx])
                    else:
                        invalid_inputs.append(item)
                else:
                    hierarchy_columns.append(item)
            
            # Проверка валидности
            invalid_columns = [col for col in hierarchy_columns if col not in common_headers]
            if invalid_columns or invalid_inputs:
                invalid_list = invalid_columns + invalid_inputs
                print(f"Error: Invalid columns: {', '.join(invalid_list)}")
                continue
            
            if not hierarchy_columns:
                print("Error: No valid columns selected")
                continue
            
            break
        
        # Шаг 3: Последовательный выбор категорий
        print("\nStarting sequential category selection...")
        all_combinations = select_categories_sequentially(source, valid_sheets, hierarchy_columns)
        
        if not all_combinations:
            print("No combinations selected")
            return False
        
        # Создание файлов
        os.makedirs(destination, exist_ok=True)
        base_name = os.path.splitext(os.path.basename(source))[0]
        created_files = []
        
        for filters in all_combinations:
            # Формируем имя файла с расширением .xlsx
            safe_parts = [sanitize_filename(v) for v in filters.values()]
            suffix = "_".join(safe_parts) if safe_parts else "All"
            target_file = os.path.join(destination, f"{base_name}_{suffix}.xlsx")
            
            # Создаем файл
            created_file = create_filtered_file(source, target_file, valid_sheets, filters)
            if created_file is not None:
                created_files.append(created_file)
        
        # Вывод результатов
        if created_files:
            print(f"\nCreated {len(created_files)} files:")
            for file in created_files:
                print(f"  - {file}")
        else:
            print("Warning: No files created (no data matched the filters)")
        
        return True
    
    except KeyboardInterrupt:
        print("\nOperation cancelled by user (Ctrl+C)")
        return False
    except Exception as e:
        print(f"Error: {str(e)}")
        return False

def main():
    """Главный цикл программы: обработка файлов."""
    while True:
        success = process_file()
        
        # Спрашиваем, хочет ли пользователь продолжить
        if success:
            cont = input("\nDo you want to process another file? (y/n): ").strip().lower()
            if cont != 'y':
                print("Program terminated by user")
                break
        else:
            cont = input("\nDo you want to try again? (y/n): ").strip().lower()
            if cont != 'y':
                print("Program terminated by user")
                break

if __name__ == "__main__":
    main()