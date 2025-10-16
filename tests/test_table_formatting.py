import unittest
import os
import tempfile
import openpyxl
from excel_utils.workbook import create_filtered_file, clean_table_name
from excel_utils.analysis import get_all_sheets_headers

class TestTableFormatting(unittest.TestCase):
    def setUp(self):
        # Создаем тестовый Excel-файл
        self.temp_dir = tempfile.mkdtemp()
        self.test_file = os.path.join(self.temp_dir, "test_table.xlsx")
        wb = openpyxl.Workbook()
        ws = wb.active
        ws.title = "TestSheet"
        
        # Добавляем технические строки
        ws.append(["Technical info 1"])
        ws.append(["Technical info 2"])
        
        # Добавляем заголовки
        ws.append(["ID", "Value", "Category"])
        
        # Добавляем данные
        for i in range(1, 11):
            ws.append([i, i*10, "A" if i % 2 == 0 else "B"])
        
        wb.save(self.test_file)
    
    def tearDown(self):
        import shutil
        shutil.rmtree(self.temp_dir)
    
    def test_table_range(self):
        """Проверяет, что таблица создается только для данных, без технических строк"""
        sheet_headers = get_all_sheets_headers(self.test_file)
        valid_sheets = {k: v for k, v in sheet_headers.items() if v[0] is not None}
        
        # Создаем простой фильтр
        filters = {}
        output_file = os.path.join(self.temp_dir, "output.xlsx")
        result = create_filtered_file(self.test_file, output_file, valid_sheets, filters)
        
        self.assertIsNotNone(result)
        self.assertTrue(os.path.exists(result))
        
        # Проверяем, что таблица создана
        wb = openpyxl.load_workbook(result)
        ws = wb["TestSheet"]
        
        # Проверяем, что таблица создана и имеет правильный диапазон
        self.assertTrue(len(ws.tables) > 0)
        table = list(ws.tables.values())[0]
        
        # Проверяем, что таблица начинается с строки заголовков (строка 3)
        self.assertEqual(table.ref[0], 'A')
        self.assertEqual(table.ref[1], '3')
        
        # Проверяем, что таблица заканчивается на последней строке данных (строка 12)
        last_row = int(table.ref.split(':')[1][1:])
        self.assertEqual(last_row, 12)
    
    def test_table_name_cleaning(self):
        """Проверяет, что имена таблиц не содержат пробелов"""
        # Проверка функции очистки имени
        self.assertEqual(clean_table_name("Sheet 1"), "Sheet1")
        self.assertEqual(clean_table_name("Таблица с пробелами"), "Таблицаспробелами")
        self.assertEqual(clean_table_name("Table@Name"), "TableName")
        self.assertEqual(clean_table_name("Table!Name"), "TableName")
        
        # Проверка, что имя таблицы не содержит пробелов
        sheet_headers = get_all_sheets_headers(self.test_file)
        valid_sheets = {k: v for k, v in sheet_headers.items() if v[0] is not None}
        
        # Создаем простой фильтр
        filters = {}
        output_file = os.path.join(self.temp_dir, "output.xlsx")
        result = create_filtered_file(self.test_file, output_file, valid_sheets, filters)
        
        self.assertIsNotNone(result)
        self.assertTrue(os.path.exists(result))
        
        # Проверяем имя таблицы
        wb = openpyxl.load_workbook(result)
        ws = wb["TestSheet"]
        
        # Проверяем, что таблица создана
        self.assertTrue(len(ws.tables) > 0)
        
        # Проверяем, что имя таблицы не содержит пробелов
        table = list(ws.tables.values())[0]
        self.assertEqual(table.name, "TableTestSheet")
    
    def test_table_with_filtering(self):
        """Проверяет создание таблицы при фильтрации данных"""
        sheet_headers = get_all_sheets_headers(self.test_file)
        valid_sheets = {k: v for k, v in sheet_headers.items() if v[0] is not None}
        
        # Создаем фильтр для категории A
        filters = {"Category": "A"}
        output_file = os.path.join(self.temp_dir, "output_filtered.xlsx")
        result = create_filtered_file(self.test_file, output_file, valid_sheets, filters)
        
        self.assertIsNotNone(result)
        self.assertTrue(os.path.exists(result))
        
        # Проверяем, что таблица создана
        wb = openpyxl.load_workbook(result)
        ws = wb["TestSheet"]
        
        # Проверяем, что таблица создана
        self.assertTrue(len(ws.tables) > 0)
        
        # Проверяем, что таблица имеет правильный диапазон
        table = list(ws.tables.values())[0]
        
        # Проверяем, что таблица начинается с строки заголовков (строка 3)
        self.assertEqual(table.ref[0], 'A')
        self.assertEqual(table.ref[1], '3')
        
        # Проверяем, что таблица заканчивается на последней строке данных (строка 8 для категории A)
        last_row = int(table.ref.split(':')[1][1:])
        self.assertEqual(last_row, 8)
    
    def test_table_with_no_extra_columns(self):
        """Проверяет, что таблица не захватывает лишние колонки"""
        sheet_headers = get_all_sheets_headers(self.test_file)
        valid_sheets = {k: v for k, v in sheet_headers.items() if v[0] is not None}
        
        # Создаем простой фильтр
        filters = {}
        output_file = os.path.join(self.temp_dir, "output.xlsx")
        result = create_filtered_file(self.test_file, output_file, valid_sheets, filters)
        
        self.assertIsNotNone(result)
        self.assertTrue(os.path.exists(result))
        
        # Проверяем, что таблица создана
        wb = openpyxl.load_workbook(result)
        ws = wb["TestSheet"]
        
        # Проверяем, что таблица создана
        self.assertTrue(len(ws.tables) > 0)
        
        # Проверяем, что таблица не захватывает лишние колонки
        table = list(ws.tables.values())[0]
        last_col = table.ref.split(':')[1][0]
        self.assertEqual(last_col, 'C')  # Последняя колонка должна быть C, так как у нас 3 колонки
    
    def test_table_with_long_name(self):
        """Проверяет создание таблицы с длинным именем листа"""
        # Создаем файл с длинным именем листа
        wb = openpyxl.Workbook()
        ws = wb.active
        ws.title = "Very Long Sheet Name With Spaces And Special Characters!"
        
        # Добавляем заголовки
        ws.append(["ID", "Value", "Category"])
        
        # Добавляем данные
        ws.append([1, 10, "A"])
        
        # Сохраняем файл
        test_file = os.path.join(self.temp_dir, "test_long_name.xlsx")
        wb.save(test_file)
        
        sheet_headers = get_all_sheets_headers(test_file)
        valid_sheets = {k: v for k, v in sheet_headers.items() if v[0] is not None}
        
        # Создаем простой фильтр
        filters = {}
        output_file = os.path.join(self.temp_dir, "output_long.xlsx")
        result = create_filtered_file(test_file, output_file, valid_sheets, filters)
        
        self.assertIsNotNone(result)
        self.assertTrue(os.path.exists(result))
        
        # Проверяем, что таблица создана
        wb = openpyxl.load_workbook(result)
        ws = wb["Very Long Sheet Name With Spaces And Special Characters!"]
        
        # Проверяем, что таблица создана
        self.assertTrue(len(ws.tables) > 0)
        
        # Проверяем имя таблицы
        table = list(ws.tables.values())[0]
        # Убедимся, что имя таблицы очищено от недопустимых символов
        self.assertNotIn(' ', table.name)
        self.assertNotIn('!', table.name)
        self.assertEqual(table.name[:5], "Table")  # Имя должно начинаться с "Table"