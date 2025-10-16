import unittest
import os
import tempfile
import openpyxl
from excel_utils.workbook import create_filtered_file
from excel_utils.analysis import get_all_sheets_headers

class TestSpecialFormatting(unittest.TestCase):
    def setUp(self):
        # Создаем тестовый Excel-файл с особыми стилями
        self.temp_dir = tempfile.mkdtemp()
        self.test_file = os.path.join(self.temp_dir, "test_special.xlsx")
        
        wb = openpyxl.Workbook()
        ws = wb.active
        ws.title = "TestSheet"
        
        # Добавляем заголовки
        ws.append(["ID", "Золотой работник", "Category", "Other"])
        
        # Добавляем данные
        ws.append([1, "да", "A", "Data"])
        ws.append([2, "нет", "B", "Data"])
        ws.append([3, "запланирован", "A", "Data"])
        ws.append([4, "пробел", "B", "Data"])
        ws.append([5, "другое значение", "A", "Data"])
        
        wb.save(self.test_file)
    
    def tearDown(self):
        import shutil
        shutil.rmtree(self.temp_dir)
    
    def test_golden_worker_formatting(self):
        """Проверяет специфическое форматирование для колонки 'Золотой работник'"""
        sheet_headers = get_all_sheets_headers(self.test_file)
        valid_sheets = {k: v for k, v in sheet_headers.items() if v[0] is not None}
        
        # Создаем простой фильтр
        filters = {}
        output_file = os.path.join(self.temp_dir, "output.xlsx")
        result = create_filtered_file(self.test_file, output_file, valid_sheets, filters)
        
        self.assertIsNotNone(result)
        self.assertTrue(os.path.exists(result))
        
        # Проверяем форматирование
        wb = openpyxl.load_workbook(result)
        ws = wb["TestSheet"]
        
        # Проверяем цвета ячеек в колонке "Золотой работник"
        # Строка 2 (данные): "да" -> зеленый
        cell = ws['B2']
        self.assertEqual(cell.fill.start_color.index, '96C850')
        
        # Строка 3: "нет" -> красный
        cell = ws['B3']
        self.assertEqual(cell.fill.start_color.index, 'FF5050')
        
        # Строка 4: "запланирован" -> красный
        cell = ws['B4']
        self.assertEqual(cell.fill.start_color.index, 'FF5050')
        
        # Строка 5: "пробел" -> красный
        cell = ws['B5']
        self.assertEqual(cell.fill.start_color.index, 'FF5050')
        
        # Строка 6: "другое значение" -> без заливки
        cell = ws['B6']
        self.assertEqual(cell.fill.start_color.index, '00000000')  # Прозрачный цвет
    
    def test_golden_worker_case_insensitivity(self):
        """Проверяет обработку регистра в названии колонки и значениях"""
        # Создаем файл с разным регистром
        wb = openpyxl.Workbook()
        ws = wb.active
        ws.title = "TestSheet"
        
        # Добавляем заголовки с разным регистром
        ws.append(["ID", "ЗОЛОТОЙ РАБОТНИК", "Category"])
        
        # Добавляем данные с разным регистром
        ws.append([1, "ДА", "A"])
        ws.append([2, "НЕТ", "B"])
        ws.append([3, "ЗАПЛАНИРОВАН", "A"])
        ws.append([4, "ПРОБЕЛ", "B"])
        
        # Сохраняем файл
        test_file = os.path.join(self.temp_dir, "test_case_insensitive.xlsx")
        wb.save(test_file)
        
        sheet_headers = get_all_sheets_headers(test_file)
        valid_sheets = {k: v for k, v in sheet_headers.items() if v[0] is not None}
        
        # Создаем простой фильтр
        filters = {}
        output_file = os.path.join(self.temp_dir, "output_case.xlsx")
        result = create_filtered_file(test_file, output_file, valid_sheets, filters)
        
        self.assertIsNotNone(result)
        self.assertTrue(os.path.exists(result))
        
        # Проверяем форматирование
        wb = openpyxl.load_workbook(result)
        ws = wb["TestSheet"]
        
        # Проверяем цвета ячеек
        # Строка 2: "ДА" -> зеленый
        cell = ws['B2']
        self.assertEqual(cell.fill.start_color.index, '96C850')
        
        # Строка 3: "НЕТ" -> красный
        cell = ws['B3']
        self.assertEqual(cell.fill.start_color.index, 'FF5050')
        
        # Строка 4: "ЗАПЛАНИРОВАН" -> красный
        cell = ws['B4']
        self.assertEqual(cell.fill.start_color.index, 'FF5050')
        
        # Строка 5: "ПРОБЕЛ" -> красный
        cell = ws['B5']
        self.assertEqual(cell.fill.start_color.index, 'FF5050')
    
    def test_golden_worker_in_middle(self):
        """Проверяет форматирование когда 'Золотой работник' не первая колонка"""
        # Создаем файл с колонкой в середине
        wb = openpyxl.Workbook()
        ws = wb.active
        ws.title = "TestSheet"
        
        # Добавляем заголовки
        ws.append(["ID", "Name", "Золотой работник", "Category"])
        
        # Добавляем данные
        ws.append([1, "John", "да", "A"])
        ws.append([2, "Jane", "нет", "B"])
        
        # Сохраняем файл
        test_file = os.path.join(self.temp_dir, "test_middle.xlsx")
        wb.save(test_file)
        
        sheet_headers = get_all_sheets_headers(test_file)
        valid_sheets = {k: v for k, v in sheet_headers.items() if v[0] is not None}
        
        # Создаем простой фильтр
        filters = {}
        output_file = os.path.join(self.temp_dir, "output_middle.xlsx")
        result = create_filtered_file(test_file, output_file, valid_sheets, filters)
        
        self.assertIsNotNone(result)
        self.assertTrue(os.path.exists(result))
        
        # Проверяем форматирование
        wb = openpyxl.load_workbook(result)
        ws = wb["TestSheet"]
        
        # Проверяем цвета ячеек
        # Строка 2: "да" -> зеленый
        cell = ws['C2']
        self.assertEqual(cell.fill.start_color.index, '96C850')
        
        # Строка 3: "нет" -> красный
        cell = ws['C3']
        self.assertEqual(cell.fill.start_color.index, 'FF5050')