from contextlib import contextmanager
import openpyxl
from .common import validate_row
import logging

logger = logging.getLogger('excel_splitter')

@contextmanager
def safe_workbook(file_path, read_only=False):
    """Контекстный менеджер для безопасной работы с файлами Excel."""
    wb = None
    try:
        wb = openpyxl.load_workbook(file_path, read_only=read_only)
        yield wb
    finally:
        if wb:
            try:
                wb.close()
            except Exception as e:
                logger.error(f"Error closing workbook: {str(e)}")

def get_all_sheets_headers(file_path, max_scan_rows=10):
    """Анализирует все ВИДИМЫЕ листы в Excel-файле, возвращает заголовки для каждого."""
    logger.info(f"Analyzing headers in {file_path}")
    try:
        with safe_workbook(file_path, read_only=True) as wb:
            sheet_results = {}
            for ws in wb.worksheets:
                # Игнорируем скрытые листы
                if ws.sheet_state != 'visible':
                    continue
                max_non_empty = 0
                header_row = None
                header_row_idx = 0
                for row_idx, row in enumerate(ws.iter_rows(min_row=1, max_row=max_scan_rows), start=1):
                    non_empty_count = sum(1 for cell in row if cell.value is not None)
                    if non_empty_count > max_non_empty:
                        max_non_empty = non_empty_count
                        header_row = row
                        header_row_idx = row_idx
                if max_non_empty > 0:
                    headers = [cell.value for cell in header_row if cell.value is not None]
                    sheet_results[ws.title] = (headers, header_row_idx)
                    logger.debug(f"Found headers in sheet {ws.title}: {headers}")
                else:
                    sheet_results[ws.title] = (None, None)
            return sheet_results
    except Exception as e:
        logger.error(f"Error analyzing Excel: {str(e)}")
        raise ValueError(f"Error analyzing Excel: {str(e)}")

def analyze_column(file_path, valid_sheets, selected_column, filters=None):
    """Собирает уникальные значения из указанной колонки с учетом фильтров."""
    if filters is None:
        filters = {}
    logger.info(f"Analyzing column {selected_column} with filters {filters}")
    try:
        with safe_workbook(file_path, read_only=True) as wb:
            categories = set()
            for sheet_name, (headers, row_idx) in valid_sheets.items():
                ws = wb[sheet_name]
                try:
                    col_index = headers.index(selected_column)
                except ValueError:
                    continue
                for row in ws.iter_rows(min_row=row_idx + 1, values_only=True):
                    if not validate_row(row, headers, row_idx, filters):
                        continue
                    cell_value = row[col_index] if col_index < len(row) else None
                    if cell_value is not None and str(cell_value).strip() != "":
                        categories.add(str(cell_value).strip())
            return sorted(categories)
    except Exception as e:
        logger.error(f"Error analyzing data: {str(e)}")
        raise ValueError(f"Error analyzing data: {str(e)}")