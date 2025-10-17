import os
import re
import logging
import openpyxl
from copy import copy
from contextlib import contextmanager
from openpyxl.worksheet.table import Table, TableStyleInfo
from excel_utils.common import validate_row, copy_cell_style
from excel_utils.formatting import sanitize_filename
from excel_utils.analysis import get_all_sheets_headers

logger = logging.getLogger('excel_splitter')

def get_column_letter(col_idx):
    """Конвертирует индекс столбца в букву (1 -> A, 26 -> Z, 27 -> AA и т.д.)"""
    letters = []
    while col_idx > 0:
        col_idx, remainder = divmod(col_idx - 1, 26)
        letters.append(chr(65 + remainder))
    return ''.join(reversed(letters))

def clean_table_name(name):
    """Очищает имя таблицы от недопустимых символов и пробелов."""
    # Удаляем недопустимые символы
    clean_name = re.sub(r'[^\w]', '', name)
    # Если имя слишком длинное, сокращаем его
    if len(clean_name) > 31:
        clean_name = clean_name[:31]
    # Если имя пустое, возвращаем дефолтное имя
    if not clean_name:
        return "Table"
    return clean_name

def is_openpyxl_new_version():
    """Проверяет, является ли версия openpyxl новой (>= 3.0)"""
    try:
        from openpyxl import __version__
        major_version = int(__version__.split('.')[0])
        return major_version >= 3
    except (ImportError, ValueError):
        return False

@contextmanager
def safe_workbook(file_path, read_only=False):
    """Контекстный менеджер для безопасной работы с файлами Excel."""
    wb = None
    try:
        logger.debug(f"Opening workbook: {file_path}")
        wb = openpyxl.load_workbook(file_path, read_only=read_only)
        yield wb
    finally:
        if wb:
            try:
                wb.close()
                logger.debug(f"Workbook closed: {file_path}")
            except Exception as e:
                logger.error(f"Error closing workbook: {str(e)}")

def copy_technical_rows(ws_source, ws_new, header_row_idx):
    """Копирует технические строки выше таблицы (строки выше заголовков)."""
    for row_idx in range(1, header_row_idx):
        for col_idx in range(1, ws_source.max_column + 1):
            try:
                source_cell = ws_source.cell(row=row_idx, column=col_idx)
                if source_cell.value is not None or source_cell.has_style:
                    target_cell = ws_new.cell(row=row_idx, column=col_idx, value=source_cell.value)
                    copy_cell_style(source_cell, target_cell)
            except Exception as e:
                logger.debug(f"Error copying cell at row {row_idx}, col {col_idx}: {str(e)}")

def copy_headers(ws_source, ws_new, header_row_idx):
    """Копирует строку заголовков."""
    for col_idx in range(1, ws_source.max_column + 1):
        try:
            source_cell = ws_source.cell(row=header_row_idx, column=col_idx)
            if source_cell.value is not None or source_cell.has_style:
                target_cell = ws_new.cell(row=header_row_idx, column=col_idx, value=source_cell.value)
                copy_cell_style(source_cell, target_cell)
        except Exception as e:
            logger.debug(f"Error copying header at col {col_idx}: {str(e)}")

def filter_data_rows(ws_source, ws_new, header_row_idx, filters, headers, sheet_name, valid_sheets):
    """Фильтрует и копирует данные в соответствии с фильтрами."""
    new_row_idx = header_row_idx + 1
    filtered_count = 0
    has_data = False
    
    for row_idx in range(header_row_idx + 1, ws_source.max_row + 1):
        try:
            row = ws_source[row_idx]
            # Добавлена проверка на пустой фильтр
            if not filters:
                should_include = True
            else:
                should_include = validate_row([cell.value for cell in row], headers, header_row_idx, filters)
            
            if should_include:
                filtered_count += 1
                has_data = True
                for col_idx in range(1, ws_source.max_column + 1):
                    try:
                        source_cell = ws_source.cell(row=row_idx, column=col_idx)
                        if source_cell.value is not None or source_cell.has_style:
                            target_cell = ws_new.cell(row=new_row_idx, column=col_idx, value=source_cell.value)
                            copy_cell_style(source_cell, target_cell)
                    except Exception as e:
                        logger.debug(f"Error copying data cell at row {row_idx}, col {col_idx}: {str(e)}")
                new_row_idx += 1
        except Exception as e:
            logger.debug(f"Error processing row {row_idx}: {str(e)}")
    
    logger.debug(f"Filtered {filtered_count} rows out of {ws_source.max_row - header_row_idx} possible")
    return has_data, new_row_idx

def determine_table_boundaries(ws_source, ws_new, header_row_idx, new_row_idx):
    """Определяет границы таблицы: последнюю колонку с данными и конечную строку."""
    last_col = 0
    for col_idx in range(1, ws_source.max_column + 1):
        # Проверяем, есть ли данные в этой колонке
        has_data_in_col = False
        for row_idx in range(header_row_idx, new_row_idx):
            if ws_new.cell(row=row_idx, column=col_idx).value is not None:
                has_data_in_col = True
                break
        if has_data_in_col:
            last_col = col_idx
    
    # Если не определили последнюю колонку, используем max_column
    if last_col == 0:
        last_col = ws_source.max_column
    
    last_col_letter = get_column_letter(last_col)
    data_start_row = header_row_idx + 1
    data_end_row = new_row_idx - 1 if new_row_idx > header_row_idx + 1 else header_row_idx
    
    return last_col_letter, data_start_row, data_end_row

def apply_table_formatting(ws_new, header_row_idx, last_col_letter, data_start_row, data_end_row):
    """Применяет форматирование таблицы к отфильтрованным данным."""
    table_range = f"A{header_row_idx}:{last_col_letter}{data_end_row}"
    # Создаем таблицу с безопасным именем
    safe_table_name = clean_table_name(ws_new.title)
    table = Table(displayName=safe_table_name, ref=table_range)
    
    # Создаем стиль таблицы с только поддерживаемыми параметрами
    try:
        # Пытаемся создать стиль с минимально необходимыми параметрами
        style = TableStyleInfo(
            name="TableStyleLight1",
            showFirstColumn=False,
            showLastColumn=False,
            showColumnHeaders=True
        )
        table.tableStyleInfo = style
    except TypeError as e:
        logger.warning(f"TableStyleInfo parameters not fully supported: {str(e)}")
        try:
            style = TableStyleInfo(
                name="TableStyleLight1",
                showColumnHeaders=True
            )
            table.tableStyleInfo = style
        except TypeError:
            logger.warning("Using minimal table style")
            style = TableStyleInfo(name="TableStyleLight1")
            table.tableStyleInfo = style
    
    ws_new.add_table(table)

def copy_worksheet_structure(ws_source, ws_new):
    """Копирует структурные элементы листа (ширина столбцов, высота строк, объединенные ячейки)."""
    # Копирование ширины столбцов
    if hasattr(ws_source, 'column_dimensions'):
        for col_letter, dim in ws_source.column_dimensions.items():
            try:
                ws_new.column_dimensions[col_letter].width = dim.width
            except Exception as e:
                logger.debug(f"Error copying column width for {col_letter}: {str(e)}")
    
    # Копирование высоты строк
    if hasattr(ws_source, 'row_dimensions'):
        for row_idx, dim in ws_source.row_dimensions.items():
            try:
                ws_new.row_dimensions[row_idx].height = dim.height
            except Exception as e:
                logger.debug(f"Error copying row height for {row_idx}: {str(e)}")
    
    # Копирование объединенных ячеек
    if hasattr(ws_source, 'merged_cells'):
        for merged_cell in ws_source.merged_cells.ranges:
            try:
                ws_new.merge_cells(str(merged_cell))
            except Exception as e:
                logger.debug(f"Error copying merged cells: {str(e)}")

def copy_conditional_formatting(ws_source, ws_new):
    """Копирует условное форматирование с исходного листа на новый."""
    if hasattr(ws_source, 'conditional_formatting'):
        for cf in ws_source.conditional_formatting:
            try:
                # Определяем, какой метод использовать для получения диапазона
                range_attr = '_get_range' if hasattr(cf, '_get_range') else 'ref'
                range_value = getattr(cf, range_attr, None) or cf._range
                
                # Определяем, какой тип правил используем
                if hasattr(cf, 'cfRule') and hasattr(cf, 'cfRules'):
                    # Новые версии openpyxl
                    for rule in cf.cfRules:
                        try:
                            ws_new.conditional_formatting.add(range_value, rule)
                        except Exception as e:
                            logger.debug(f"Error adding rule in new format: {str(e)}")
                elif hasattr(cf, 'rules'):
                    # Средние версии
                    for rule in cf.rules:
                        try:
                            ws_new.conditional_formatting.add(range_value, rule)
                        except Exception as e:
                            logger.debug(f"Error adding rule in medium format: {str(e)}")
                else:
                    # Старые версии
                    try:
                        ws_new.conditional_formatting.add(range_value, cf)
                    except Exception as e:
                        logger.debug(f"Error adding rule in old format: {str(e)}")
            except Exception as e:
                logger.debug(f"Error copying conditional formatting: {str(e)}")

def create_filtered_file(source, target, valid_sheets, filters):
    """Создаёт файл с фильтрацией по комбинации условий."""
    logger.info(f"Creating filtered file: {target} with filters {filters}")
    # Добавлена проверка на пустой фильтр
    if not filters:
        logger.info("Empty filters, copying all data")
    try:
        # Всегда сохраняем как .xlsx
        if target.lower().endswith('.xlsm'):
            logger.debug("Converting .xlsm to .xlsx format")
            target = target[:-5] + '.xlsx'
        with safe_workbook(source, read_only=False) as wb_source:
            wb_new = openpyxl.Workbook()
            wb_new.remove(wb_new.active)
            has_data = False  # Флаг наличия данных
            logger.debug(f"Processing {len(wb_source.sheetnames)} sheets")
            for sheet_name in wb_source.sheetnames:
                ws_source = wb_source[sheet_name]
                # Игнорируем скрытые листы
                if ws_source.sheet_state != 'visible':
                    logger.debug(f"Skipping hidden sheet: {sheet_name}")
                    continue
                ws_new = wb_new.create_sheet(title=sheet_name)
                logger.debug(f"Processing sheet: {sheet_name}")
                
                # Копируем структурные элементы листа
                copy_worksheet_structure(ws_source, ws_new)
                
                # Копируем условное форматирование
                copy_conditional_formatting(ws_source, ws_new)
                
                if sheet_name in valid_sheets:
                    headers, header_row_idx = valid_sheets[sheet_name]
                    logger.debug(f"Headers for sheet {sheet_name}: {headers}")
                    logger.debug(f"Header row index: {header_row_idx}")
                    
                    # 1. Технические строки выше таблицы
                    copy_technical_rows(ws_source, ws_new, header_row_idx)
                    
                    # 2. Заголовки
                    copy_headers(ws_source, ws_new, header_row_idx)
                    
                    # 3. Фильтрация данных
                    sheet_has_data, new_row_idx = filter_data_rows(
                        ws_source, ws_new, header_row_idx, filters, 
                        headers, sheet_name, valid_sheets
                    )
                    
                    if sheet_has_data:
                        has_data = True
                        # Определяем границы таблицы
                        last_col_letter, data_start_row, data_end_row = determine_table_boundaries(
                            ws_source, ws_new, header_row_idx, new_row_idx
                        )
                        
                        # Применяем форматирование таблицы
                        apply_table_formatting(
                            ws_new, header_row_idx, last_col_letter, 
                            data_start_row, data_end_row
                        )
                    else:
                        # Удаляем лист без данных
                        wb_new.remove(ws_new)
                        logger.debug(f"Removed sheet {sheet_name} due to no matching data")
                        continue  # Переходим к следующему листу
                else:
                    logger.debug(f"Copying entire sheet {sheet_name} without filtering")
                    for row_idx in range(1, ws_source.max_row + 1):
                        for col_idx in range(1, ws_source.max_column + 1):
                            try:
                                source_cell = ws_source.cell(row=row_idx, column=col_idx)
                                if source_cell.value is not None or source_cell.has_style:
                                    target_cell = ws_new.cell(row=row_idx, column=col_idx, value=source_cell.value)
                                    copy_cell_style(source_cell, target_cell)
                            except Exception as e:
                                logger.debug(f"Error copying cell at row {row_idx}, col {col_idx}: {str(e)}")
            
            if not has_data:
                logger.warning("No data matched the filters, file not created")
                return None
            
            # Удаляем целевой файл, если он существует
            if os.path.exists(target):
                logger.info(f"Removing existing target file: {target}")
                os.remove(target)
            
            # Сохраняем как .xlsx
            logger.info(f"Saving filtered file: {target}")
            wb_new.save(target)
            return target
    except Exception as e:
        logger.exception(f"Error during filtering: {str(e)}")
        raise ValueError(f"Error during filtering: {str(e)}")